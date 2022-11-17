from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import string
import numpy as np
import datetime
from imap_tools import MailBox, AND
from xls2xlsx import XLS2XLSX
import os
from openpyxl import load_workbook
import pandas as pd
import tempfile
import time
# pd.options.mode.chained_assignment = None  # default='warn'
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
pd.options.mode.chained_assignment = None
from config import EMAIL,PASSWORD
import random


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller
    Функция чтобы логотип отображался"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)



def select_end_folder():
    """
    Функция для выбора конечной папки куда будут складываться итоговые файлы
    :return:
    """
    global path_to_end
    path_to_end = filedialog.askdirectory()


def getMergedCellVal(sheet, cell):
    """
    Функция для получения значения объединеной ячейки
    Взято отсюда https://stackoverflow.com/questions/23562366/how-to-get-value-present-in-a-merged-cell
    """
    rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
    return sheet.cell(rng[0].min_row, rng[0].min_col).value if len(rng)!=0 else cell.value

def fix_inn_bur(inn):
    """
    Функция для исправления инн Бурятии,для случаев когда лидирующий ноль съедается экселем
        """
    inn = inn.replace('.', '').strip()  # Удаляем точку и пробелы
    inn_len = len(inn)
    if inn_len == 10:
        return inn
    elif inn_len == 9 and inn.startswith('3'):
        return f'0{inn}'
    else:
        return f'ИНН юридического лица состоит из 10 цифр! - {inn}'


def processing_data():
    """
    Фугкция для обработки данных
    :return:
    """
    not_used = ['Спам', 'Отправленные', 'Черновики', 'Корзина']
    cols_df = list(range(24))
    df = pd.DataFrame(columns=cols_df)  # базовый датафрейм
    df['Тип таблицы'] = None
    us_df = pd.DataFrame(columns=['Откуда прислан файл', 'Название файла', 'Время отправки',
                                  'Тип ошибки'])  # Датафрейм для неправильных файлов
    dir_files_org = 'Присланные формы ФГИС по организациям'  # название папки куда будут сохраняться скачанные формы
    dir_files_other_excel = 'Файлы Excel не соответствующие форме'
    dir_other_files = 'Файлы с другими форматами'
    if not os.path.exists(f'{path_to_end}/{dir_files_org}'):  # проверяем наличие папки
        os.makedirs(f'{path_to_end}/{dir_files_org}')  # если ее нет то создаем

    if not os.path.exists(f'{path_to_end}/{dir_files_other_excel}'):  # проверяем наличие папки
        os.makedirs(f'{path_to_end}/{dir_files_other_excel}')  # если ее нет то создаем

    if not os.path.exists(f'{path_to_end}/{dir_other_files}'):  # проверяем наличие папки
        os.makedirs(f'{path_to_end}/{dir_other_files}')  # если ее нет то создаем

    with tempfile.TemporaryDirectory() as temp_dir:
        with MailBox('imap.mail.ru').login(EMAIL, PASSWORD) as mailbox:
            for f in mailbox.folder.list():
                if f.name not in not_used:
                    mailbox.folder.set(f.name)
                    for msg in mailbox.fetch():
                        # print(f' Subject {msg.subject}') # заголовок письма
                        # print(f' From {msg.from_}') # адрес почты отправителя
                        # print(f' Date {msg.date}') # время отправки
                        #
                        # print('***************')
                        msg_from = msg.from_  # получаем адрес почты отправителя
                        msg_date = msg.date
                        for att in msg.attachments:
                            try:  # оборачиваем в try except чтобы при ошибках процесс продолжался
                                extension = att.filename.split('.')[
                                    -1].lower()  # получаем расширение файла и делаем его строчным(для случаев наподобие XLSX)
                                print(msg_from)
                                print(att.filename)
                                if extension == 'xlsx' or extension == 'xls':  # проверяем на расширение

                                    if extension == 'xlsx':  # Сохраняем во временную папку
                                        work_file_name = att.filename
                                        with open(f'{temp_dir}{att.filename}', 'wb') as f:
                                            f.write(att.payload)
                                    elif extension == 'xls':  # конвертируем и сохраняем
                                        work_file_name = att.filename.replace('.xls', '.xlsx')
                                        with open(f'{temp_dir}{att.filename}', 'wb') as f:
                                            f.write(att.payload)
                                        out = XLS2XLSX(f'{temp_dir}{att.filename}')  # конвертируем в xlsx
                                        out.to_xlsx((f'{temp_dir}{work_file_name}'))  # сохраняем
                                        os.remove(f'{temp_dir}{att.filename}')  # удаляем файл xls чтобы не мешался

                                    wb = load_workbook(f'{temp_dir}{work_file_name}')

                                    first_list = wb.sheetnames[0]  # получаем первый лист
                                    standard_str = 'На обработку моих персональных данных в целях подключения к Личному кабинету в gosuslugi.ru:'  # проверочная строка

                                    check_file = getMergedCellVal(wb[first_list], wb[first_list][
                                        'L2'])  # получаем значение ячейки,если совпадает то файл нужный нам
                                    if check_file == standard_str:
                                        if len(wb.sheetnames) == 1:  # Проверяем длину
                                            name_sheet = wb.sheetnames[0]  # получаем название листа
                                            name_org = wb[first_list]['B5'].value  # получаем значение ячейки B5
                                            print(name_org)
                                            print('*******')
                                            temp_df = pd.read_excel(f'{temp_dir}{work_file_name}', skiprows=4,
                                                                    header=None,
                                                                    dtype=str)  # считываем датафрейм
                                            temp_df.dropna(thresh=15, inplace=True)
                                            temp_df[0] = msg_from  # добавляем от кого
                                            temp_df.insert(1, 'Тип таблицы', name_sheet)  # добавляем название листа
                                            temp_df.insert(2, 'Время отправки',
                                                           msg.date)  # добавляем дату и время отправления
                                            temp_df['Время отправки'] = temp_df['Время отправки'].apply(
                                                lambda x: pd.to_datetime(x))
                                        else:
                                            len_sheets = len(wb.sheetnames)
                                            temp_df = pd.DataFrame(columns=list(range(24)))
                                            for sheet in wb.sheetnames:
                                                ml_temp_df = pd.read_excel(f'{temp_dir}{work_file_name}',
                                                                           sheet_name=sheet,
                                                                           skiprows=4, header=None, dtype=str)

                                                try:
                                                    check_cols = ml_temp_df.iloc[:,
                                                                 1].any()  # если есть хоть одно значение в колоноке 1 то добавляем эти данные
                                                    if check_cols:
                                                        name_org = ml_temp_df.iloc[0, 1]
                                                        print(name_org)
                                                        print('**********')
                                                        name_sheet = sheet
                                                        ml_temp_df.dropna(thresh=15, inplace=True)
                                                        ml_temp_df[0] = msg_from
                                                        ml_temp_df.insert(1, 'Тип таблицы', name_sheet)
                                                        ml_temp_df.insert(2, 'Время отправки',
                                                                          msg_date)  # добавляем дату и время отправления
                                                        ml_temp_df['Время отправки'] = ml_temp_df[
                                                            'Время отправки'].apply(
                                                            lambda x: pd.to_datetime(x))
                                                        temp_df = pd.concat([temp_df, ml_temp_df], ignore_index=True)

                                                except IndexError:
                                                    continue

                                        df = pd.concat([df, temp_df], ignore_index=True)

                                        if name_org:  # Сохраняем файл если есть имя организации
                                            name_org = name_org.translate(str.maketrans('', '',
                                                                                        string.punctuation))  # удаляем знаки препинания,которые могут помешать сохранить файлы
                                            name_org = re.sub(r'\n', ' ', name_org)  # очищаем от  символов новой строки
                                            name_org = re.sub(r'^\s+|\t|\s+$', '',
                                                              name_org)  # и табов,пробелов в начале и конце

                                            wb.save(
                                                f'{path_to_end}/{dir_files_org}/{name_org}.xlsx')  # Сохраняем файл под названием организации
                                        else:  # если не заполнено то сохраняем под емайлом откуда прислан файл.
                                            temp_bad = pd.DataFrame(
                                                columns=['Откуда прислан файл', 'Название файла', 'Время отправки',
                                                         'Тип ошибки'],
                                                data=[[
                                                    msg_from, att.filename, msg_date,
                                                    'Незаполненный файл !!!']])  # создаем датафрейм с данными ошибки
                                            temp_bad['Время отправки'] = temp_bad['Время отправки'].apply(
                                                lambda x: pd.to_datetime(x))
                                            us_df = pd.concat([us_df, temp_bad],
                                                              ignore_index=True)  # добавляем в список ошибок

                                            wb.save(f'{path_to_end}/{dir_files_org}/{msg_from}.xlsx')
                                    else:
                                        # Если файл Excel не подходит под форму то сохраняем его в отдельную папку
                                        wb.save(f'{path_to_end}/{dir_files_other_excel}/{msg_from}_{work_file_name}')

                                else:
                                    with open(f'{path_to_end}/{dir_other_files}/{msg_from}_{att.filename}', 'wb') as f:
                                        f.write(att.payload)

                                    data = [msg_from, att.filename, msg_date, 'Неправильный формат !!!']

                                    temp_bad = pd.DataFrame(
                                        columns=['Откуда прислан файл', 'Название файла', 'Время отправки',
                                                 'Тип ошибки'],
                                        data=[data])  # создаем датафрейм с данными ошибки
                                    temp_bad['Время отправки'] = temp_bad['Время отправки'].apply(
                                        lambda x: pd.to_datetime(x))

                                    us_df = pd.concat([us_df, temp_bad], ignore_index=True)
                            except:

                                temp_bad = pd.DataFrame(
                                    columns=['Откуда прислан файл', 'Название файла', 'Время отправки', 'Тип ошибки'],
                                    data=[[
                                        msg_from, att.filename, msg_date,
                                        'Ошибка при обработке файла !!!']])  # создаем датафрейм с данными ошибки
                                temp_bad['Время отправки'] = temp_bad['Время отправки'].apply(
                                    lambda x: pd.to_datetime(x))
                                us_df = pd.concat([us_df, temp_bad], ignore_index=True)
                                continue

    t = time.localtime()
    current_time = time.strftime('%H_%M_%S', t)

    df.rename(columns={0: 'Откуда прислан файл', 1: 'Название учреждения'}, inplace=True)
    df.insert(1, 'Лист', df['Тип таблицы'])
    df['Время отправки'] = df['Время отправки'].apply(
        lambda a: datetime.datetime.strftime(a, "%Y-%m-%d %H:%M:%S"))  # удаляем таймзону конвертируя в строку
    df['Время отправки'] = pd.to_datetime(df['Время отправки'])  # конвертируем обратно в дату
    df.sort_values(by='Время отправки', inplace=True)  # сортируем по времени

    df.insert(2, 'Дата и время отправки', df['Время отправки'])
    # df.drop(columns=['Тип таблицы', 'Время отправки', 23], inplace=True)

    df.rename(columns={2: 'Тип', 3: 'Наименование', 4: 'Краткое наименование населенного пункта',
                       5: 'Наименование муниципального района, муниципального/городского округа',
                       6: 'Регион', 7: 'ИНН', 8: 'ОГРН', 9: 'Email', 10: 'Телефон', 11: 'Согласие директора',
                       12: 'ФИО директора', 13: 'Должность директора',
                       14: 'Телефон директора', 15: 'СНИЛС директора', 16: 'Email директора',
                       17: 'Согласие администратора',
                       18: 'ФИО администратора',
                       19: 'Должность администратора', 20: 'Телефон администратора', 21: 'СНИЛС администратора',
                       22: 'Email администратора'}, inplace=True)

    df.drop(columns=df.iloc[:, 25:], inplace=True)

    df['Название учреждения'] = df['Название учреждения'].replace('', np.nan)
    df['Название учреждения'] = df['Название учреждения'].fillna(
        f'{random.random()}')  # заполняем рандомным числом, чтобы при очистке от дубликатов не удалилось

    df.drop_duplicates(subset=['Название учреждения'], keep='last', inplace=True)  # удаляем дубликаты

    df['ИНН'] = df['ИНН'].apply(fix_inn_bur)

    df.to_excel(f'{path_to_end}/Данные организаций для ФГИС Моя Школа от {current_time}.xlsx', index=False)

    us_df['Название файла'] = us_df['Название файла'].replace('', np.nan)

    us_df.dropna(inplace=True)
    us_df['Время отправки'] = us_df['Время отправки'].apply(
        lambda a: datetime.datetime.strftime(a, "%Y-%m-%d %H:%M:%S"))  # удаляем таймзону конвертируя в строку
    us_df['Время отправки'] = pd.to_datetime(us_df['Время отправки'])  # конвертируем обратно в дату
    us_df.sort_values(by='Время отправки', inplace=True)  # сортируем по времени

    us_df = pd.merge(us_df, df, how='outer', left_on='Откуда прислан файл', right_on='Откуда прислан файл',
                     indicator=True)  # мерджим файлы
    out_error_df = us_df[us_df['_merge'] != 'right_only']  # отбираем только те котороые есть и основном и ошибочном
    out_error_df.drop_duplicates(subset=['Откуда прислан файл'], keep='last',
                                 inplace=True)  # убираем дубликаты оставляя только последний присланнный неправильый файл

    out_error_df = out_error_df.drop(columns=out_error_df.iloc[:, 4:-1], axis=1)  # удаляем лишние столбцы
    out_error_df.rename(columns={'_merge': 'Итоговый результат'}, inplace=True)  # переименовываем колонку
    out_error_df['Итоговый результат'] = out_error_df['Итоговый результат'].apply(lambda x:
                                                                                  'Данные добавлены в основую таблицу из сопутствующего файла Excel' if x == 'both'
                                                                                  else 'Отсутствуют в основной таблице. Прислан пустой файл формы или файл формы не в формате Excel ')

    out_error_df.to_excel(f'{path_to_end}/Ошибки и некорректные файлы для ФГИС Моя Школа от {current_time}.xlsx',
                          index=False)




    messagebox.showinfo(message='Обработка завершена! ')



if __name__ == '__main__':
    window = Tk()
    window.title('Dodger ver 1.3')
    window.geometry('700x560')
    window.resizable(False, False)


    # Создаем объект вкладок

    tab_control = ttk.Notebook(window)

    # Создаем вкладку обработки данных для Приложения 6
    tab_report_6 = ttk.Frame(tab_control)
    tab_control.add(tab_report_6, text='Скрипт №1')
    tab_control.pack(expand=1, fill='both')
    # Добавляем виджеты на вкладку Создание образовательных программ
    # Создаем метку для описания назначения программы
    lbl_hello = Label(tab_report_6,
                      text='Центр опережающей профессиональной подготовки Республики Бурятия\n'
                           'Скрипт для получения данных из электронной почты ФГИС Моя школа')
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img = resource_path('logo.png')

    img = PhotoImage(file=path_to_img)
    Label(tab_report_6,
          image=img
          ).grid(column=1, row=0, padx=10, pady=25)

    btn_choose_end_folder = Button(tab_report_6, text='1) Выберите конечную папку', font=('Arial Bold', 20),
                                       command=select_end_folder
                                       )
    btn_choose_end_folder.grid(column=0, row=3, padx=10, pady=10)

    #Создаем кнопку обработки данных

    btn_proccessing_data = Button(tab_report_6, text='2) Получить данные', font=('Arial Bold', 20),
                                       command=processing_data
                                       )
    btn_proccessing_data.grid(column=0, row=4, padx=10, pady=10)

    window.mainloop()