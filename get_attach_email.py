"""
Получение вложений из почты для дальнейшей обработки

"""
import string

from imap_tools import MailBox, AND
from xls2xlsx import XLS2XLSX
import os
from openpyxl import load_workbook
import pandas as pd
import tempfile
import time
import re
import datetime
from config import EMAIL,PASSWORD

def getMergedCellVal(sheet, cell):
    """
    Функция для получения значения объединеной ячейки
    Взято отсюда https://stackoverflow.com/questions/23562366/how-to-get-value-present-in-a-merged-cell
    """
    rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
    return sheet.cell(rng[0].min_row, rng[0].min_col).value if len(rng)!=0 else cell.value



path_to_end = 'C:/Данные'
# Get date, subject and body len of all emails from INBOX folder



not_used = ['Спам','Отправленные','Черновики','Корзина']
cols_df = list(range(23))
df = pd.DataFrame(columns=cols_df) # базовый датафрейм
df['Тип таблицы'] = None
us_df = pd.DataFrame(columns=['Откуда прислан файл','Название файла','Время отправки','Тип ошибки']) # Датафрейм для неправильных файлов
dir_files_org = 'Присланные формы ФГИС по организациям' # название папки куда будут сохраняться скачанные формы
dir_files_other_excel = 'Файлы Excel не соответствующие форме'
dir_other_files = 'Файлы с другими форматами'
if not os.path.exists(f'{path_to_end}/{dir_files_org}'): # проверяем наличие папки
    os.makedirs(f'{path_to_end}/{dir_files_org}') # если ее нет то создаем

if not os.path.exists(f'{path_to_end}/{dir_files_other_excel}'): # проверяем наличие папки
    os.makedirs(f'{path_to_end}/{dir_files_other_excel}') # если ее нет то создаем

if not os.path.exists(f'{path_to_end}/{dir_other_files}'):  # проверяем наличие папки
    os.makedirs(f'{path_to_end}/{dir_other_files}')  # если ее нет то создаем


with tempfile.TemporaryDirectory() as temp_dir:


    with MailBox('imap.mail.ru').login(EMAIL,PASSWORD) as mailbox:
        for f in mailbox.folder.list():
            if f.name not in not_used:
                mailbox.folder.set(f.name)
                for msg in mailbox.fetch():
                    # print(f' Subject {msg.subject}') # заголовок письма
                    # print(f' From {msg.from_}') # адрес почты отправителя
                    # print(f' Date {msg.date}') # время отправки
                    #
                    # print('***************')
                    msg_from = msg.from_ # получаем адрес почты отправителя
                    msg_date = msg.date
                    for att in msg.attachments:
                        try: # оборачиваем в try except чтобы при ошибках процесс продолжался
                            extension = att.filename.split('.')[-1].lower() # получаем расширение файла и делаем его строчным(для случаев наподобие XLSX)
                            print(msg_from)
                            print(att.filename)
                            if extension == 'xlsx' or extension == 'xls':  # проверяем на расширение

                                if extension == 'xlsx': # Сохраняем во временную папку
                                    work_file_name = att.filename
                                    with open(f'{temp_dir}{att.filename}','wb') as f:
                                        f.write(att.payload)
                                elif extension == 'xls': # конвертируем и сохраняем
                                    work_file_name = att.filename.replace('.xls', '.xlsx')
                                    with open(f'{temp_dir}{att.filename}', 'wb') as f:
                                        f.write(att.payload)
                                    out = XLS2XLSX(f'{temp_dir}{att.filename}')  # конвертируем в xlsx
                                    out.to_xlsx((f'{temp_dir}{work_file_name}'))  # сохраняем
                                    os.remove(f'{temp_dir}{att.filename}')  # удаляем файл xls чтобы не мешался

                                wb = load_workbook(f'{temp_dir}{work_file_name}')

                                first_list = wb.sheetnames[0] # получаем первый лист
                                standard_str = 'На обработку моих персональных данных в целях подключения к Личному кабинету в gosuslugi.ru:' # проверочная строка

                                check_file = getMergedCellVal(wb[first_list], wb[first_list]['L2']) # получаем значение ячейки,если совпадает то файл нужный нам
                                if check_file == standard_str:
                                    if len(wb.sheetnames) == 1: # Проверяем длину
                                        name_sheet = wb.sheetnames[0] # получаем название листа
                                        name_org = wb[first_list]['B5'].value # получаем значение ячейки B5
                                        print(name_org)
                                        print('*******')
                                        temp_df = pd.read_excel(f'{temp_dir}{work_file_name}',skiprows=4,header=None,dtype=str) # считываем датафрейм
                                        temp_df.dropna(thresh=15,inplace=True)
                                        temp_df[0] = msg_from
                                        temp_df.insert(1,'Тип таблицы',name_sheet)
                                    else:
                                        len_sheets = len(wb.sheetnames)
                                        temp_df = pd.DataFrame(columns=list(range(23)))
                                        for sheet in wb.sheetnames:
                                            ml_temp_df = pd.read_excel(f'{temp_dir}{work_file_name}',sheet_name=sheet,skiprows=4,header=None,dtype=str)

                                            try:
                                                check_cols = ml_temp_df.iloc[:,1].any() # если есть хоть одно значение в колоноке 1 то добавляем эти данные
                                                if check_cols:
                                                    name_org = ml_temp_df.iloc[0,1]
                                                    print(name_org)
                                                    print('**********')
                                                    name_sheet = sheet
                                                    ml_temp_df.dropna(thresh=15, inplace=True)
                                                    ml_temp_df[0] = msg_from
                                                    ml_temp_df.insert(1, 'Тип таблицы', name_sheet)
                                                    temp_df=pd.concat([temp_df,ml_temp_df],ignore_index=True)

                                            except IndexError:
                                                continue

                                    df = pd.concat([df,temp_df],ignore_index=True)

                                    if name_org: # Сохраняем файл если есть имя организации
                                        name_org = name_org.translate(str.maketrans('','',string.punctuation)) # удаляем знаки препинания,которые могут помешать сохранить файлы
                                        name_org = re.sub(r'\n', ' ', name_org)# очищаем от  символов новой строки
                                        name_org = re.sub(r'^\s+|\t|\s+$', '', name_org)#  и табов,пробелов в начале и конце

                                        wb.save(f'{path_to_end}/{dir_files_org}/{name_org}.xlsx') # Сохраняем файл под названием организации
                                    else: # если не заполнено то сохраняем под емайлом откуда прислан файл.
                                        wb.save(f'{path_to_end}/{dir_files_org}/{msg_from}.xlsx')
                                else:
                                    wb.save(f'{path_to_end}/{dir_files_other_excel}/{msg_from}_{work_file_name}')

                            else:
                                with open(f'{path_to_end}/{dir_other_files}/{msg_from}_{att.filename}', 'wb') as f:
                                    f.write(att.payload)

                                data = [msg_from,att.filename,msg_date,'Неправильный формат !!!']

                                temp_bad = pd.DataFrame(columns=['Откуда прислан файл','Название файла','Время отправки','Тип ошибки'],data=[data]) # создаем датафрейм с данными ошибки
                                temp_bad['Время отправки'] = temp_bad['Время отправки'].apply(lambda x: pd.to_datetime(x).date())

                                us_df = pd.concat([us_df,temp_bad],ignore_index=True)
                        except:

                            temp_bad = pd.DataFrame(columns=['Откуда прислан файл','Название файла','Время отправки','Тип ошибки'],
                                                    data=(
                                                    msg_from, att.filename, msg_date, 'Ошибка при обработке файла !!!'))  # создаем датафрейм с данными ошибки
                            temp_bad['Время отправки'] = temp_bad['Время отправки'].apply(lambda x: pd.to_datetime(x).date())
                            us_df = pd.concat([us_df,temp_bad], ignore_index=True)
                            continue

t = time.localtime()
current_time = time.strftime('%H_%M_%S', t)

df.rename(columns={0:'Откуда прислан файл',1:'Название учреждения'},inplace=True)
df.to_excel('Тесе.xlsx',index=False)
df.insert(1,'Лист',df['Тип таблицы'])
df.drop(columns=['Тип таблицы'],inplace=True)

df.rename(columns={2:'Тип',3:'Наименование',4:'Краткое наименование населенного пункта',5:'Наименование муниципального района, муниципального/городского округа',
                   6:'Регион',7:'ИНН',8:'ОГРН',9:'Email',10:'Телефон',11:'Согласие директора',12:'ФИО директора',13:'Должность директора',
                   14:'Телефон директора',15:'СНИЛС директора',16:'Email директора',17:'Согласие администратора',18:'ФИО администратора',
                   19:'Должность администратора',20:'Телефон администратора',21:'СНИЛС администратора',22:'Email администратора'},inplace=True)
df.to_excel(f'{path_to_end}/Данные организаций для ФГИС Моя Школа от {current_time}.xlsx',index=False)

us_df.to_excel(f'{path_to_end}/Ошибки и некорректные файлы для ФГИС Моя Школа от {current_time}.xlsx',index=False)
